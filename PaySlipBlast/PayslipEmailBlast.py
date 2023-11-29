import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from PyPDF2 import PdfMerger


class PayslipEmailBlast:
    def run_download(self):
    # Define the scope and credentials
        scope = ['https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive']

        # Replace 'credentials.json' with your generated JSON file from Google Cloud Platform
        credentials = ServiceAccountCredentials.from_json_keyfile_name('downloadgsheet-406602-1c88b7cdcfc8.json', scope)

        # Authorize the client
        client = gspread.authorize(credentials)

        # Open the desired Google Sheet by its URL or title
        sheet_url = 'DownloadThis'
        spreadsheet = client.open(sheet_url)

        # Create an ExcelWriter object
        excel_file_name = 'Payroll.xlsx'
        excel_writer = pd.ExcelWriter(excel_file_name, engine='xlsxwriter')

        for sheet in spreadsheet.worksheets():
            
            # Get all values from the sheet
            data = sheet.get_all_values()

            # Convert the data to a Pandas DataFrame
            df = pd.DataFrame(data)

            # Remove empty rows and columns
            df = df.loc[:, ~(df == '').all(axis=0)]
            df = df.loc[~(df == '').all(axis=1)]

            # Write each DataFrame to the Excel file as a separate sheet
            df.to_excel(excel_writer, sheet_name=sheet.title, index=False, header=False)

        # Save and close the ExcelWriter
        excel_writer.close()

    def convert_sheetPdf(self, ExcelFile):
        # Replace 'your_excel_file.xlsx' with the path to your Excel file
        file_path = 'Payroll.xlsx'

        # Load the workbook
        workbook = load_workbook(file_path)

        # Initialize PDF merger
        pdf_merger = PdfMerger()

        # # Display the sheet names
        # print("Sheet names in the Excel file:")
        # for sheet_name in sheet_names:
        #     print(sheet_name)

        # Iterate through each sheet in the workbook and save it as a separate PDF
        for sheet_name in workbook.sheetnames:
            # Create a PDF filename based on the sheet name
            output_pdf = f"{sheet_name}.pdf"
            
            # Create a BytesIO object to temporarily store PDF data
            pdf_bytes = openpyxl.worksheet.export_to_pdf(workbook[sheet_name])
            
            # Add the PDF data to the PDF merger
            pdf_merger.append(pdf_bytes, bookmark=sheet_name)
